from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

class ExcelHandler:
    """Handles import/export of data to/from Excel files"""
    
    def export(self, filename, materials, products, compositions, cost_calculator):
        """Export all data to an Excel file"""
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Materials sheet
        ws_mat = wb.create_sheet("Materie Prime")
        self._setup_materials_sheet(ws_mat, materials)
        
        # Products sheet
        ws_prod = wb.create_sheet("Prodotti")
        self._setup_products_sheet(ws_prod, products, cost_calculator)
        
        # Compositions sheet
        ws_comp = wb.create_sheet("Composizioni")
        self._setup_compositions_sheet(ws_comp, compositions, materials, products)
        
        wb.save(filename)
    
    def _setup_materials_sheet(self, ws, materials):
        """Setup and fill materials sheet"""
        headers = ['ID', 'Nome', 'Unità', 'Costo unitario']
        ws.append(headers)
        
        for header in headers:
            ws[f'{chr(64 + headers.index(header) + 1)}1'].font = Font(bold=True)
            ws[f'{chr(64 + headers.index(header) + 1)}1'].fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        
        for material in materials:
            ws.append([material.id, material.name, material.unit, material.cost_per_unit])
        
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 15
    
    def _setup_products_sheet(self, ws, products, cost_calculator):
        """Setup and fill products sheet"""
        headers = ['ID', 'Nome', 'Semilavorato', 'Costo calcolato']
        ws.append(headers)
        
        for header in headers:
            ws[f'{chr(64 + headers.index(header) + 1)}1'].font = Font(bold=True)
            ws[f'{chr(64 + headers.index(header) + 1)}1'].fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        
        for product in products:
            cost = cost_calculator.calculate_product_cost(product.id)
            ws.append([product.id, product.name, product.is_semilavorato, cost])
        
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
    
    def _setup_compositions_sheet(self, ws, compositions, materials, products):
        """Setup and fill compositions sheet"""
        headers = ['ID Prodotto', 'Nome Prodotto', 'ID Componente', 'Nome Componente', 'Quantità', 'Tipo']
        ws.append(headers)
        
        for header in headers:
            col_idx = headers.index(header) + 1
            ws[f'{chr(64 + col_idx)}1'].font = Font(bold=True)
            ws[f'{chr(64 + col_idx)}1'].fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        
        for composition in compositions.values():
            product = products.get(composition.product_id)
            for component_id, quantity, component_type in composition.get_components():
                if component_type == "raw_material":
                    material = materials.get(component_id)
                    component_name = material.name if material else "Unknown"
                else:
                    prod = products.get(component_id)
                    component_name = prod.name if prod else "Unknown"
                
                ws.append([
                    composition.product_id,
                    product.name if product else "Unknown",
                    component_id,
                    component_name,
                    quantity,
                    component_type
                ])
        
        for i in range(1, 7):
            ws.column_dimensions[chr(64 + i)].width = 20
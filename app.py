# ============================================================================
# INCLUDE
# ============================================================================
import sys
import sqlite3
from pathlib import Path
from PyQt5.QtWidgets import (QApplication, QMainWindow, QTabWidget, QWidget, QVBoxLayout, QHBoxLayout,
                             QTableWidget, QTableWidgetItem, QComboBox, QStyledItemDelegate,
                             QLineEdit, QPushButton, QFileDialog, QMessageBox, QLabel, QDoubleSpinBox)
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QColor, QBrush
from openpyxl import Workbook, load_workbook

# ============================================================================
# VARIABILI
# ============================================================================
NOME_TAB_1 = "Materie Prime"
NOME_TAB_2 = "Prodotti"
NOME_TAB_3 = "Costi"

COLONNA_NOME = "Nome"
COLONNA_COSTO = "Costo Confezione (€)"
COLONNA_QUANTITA = "Quantità"
COLONNA_UNITA = "Unità di Misura"

UNITA_MISURA = ["kg", "grammi", "litri", "ml", "pezzo/i", "m", "cm"]

TIPI_PRODOTTO = ["Candela soia", "Jesmonite", "Candela soia + involucro jesmonite", "Lampada"]

# ============================================================================
# COSTANTI
# ============================================================================
WINDOW_WIDTH = 1400
WINDOW_HEIGHT = 700
WINDOW_TITLE = "App Candele"

FATTORI_CONVERSIONE = {
    "kg": 1.0,
    "grammi": 1000.0,
    "litri": 1.0,
    "ml": 1000.0,
    "m": 1.0,
    "cm": 100.0,
    "pezzo/i": 1.0
}

PROJECT_DIR = Path(__file__).parent.absolute()
DATA_DIR = PROJECT_DIR / "data"
DB_PATH = DATA_DIR / "materie_prime.db"

# ============================================================================
# FUNZIONI E CLASSI
# ============================================================================

def parse_float(value):
    """Converte stringa a float, accetta virgola e punto"""
    try:
        return float(str(value).replace(",", "."))
    except:
        return 0.0

def init_db():
    """Inizializza il database"""
    DATA_DIR.mkdir(exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    
    c.execute('''CREATE TABLE IF NOT EXISTS materie_prime
                 (id INTEGER PRIMARY KEY,
                  nome TEXT,
                  costo REAL,
                  quantita REAL,
                  unita TEXT)''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS prodotti
                 (id INTEGER PRIMARY KEY,
                  nome TEXT,
                  tipo TEXT)''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS composizione
                 (id INTEGER PRIMARY KEY,
                  prodotto_id INTEGER,
                  elemento_id INTEGER,
                  elemento_tipo TEXT,
                  quantita REAL,
                  unita TEXT,
                  FOREIGN KEY(prodotto_id) REFERENCES prodotti(id))''')
    
    # Aggiungi la colonna unita se non esiste
    try:
        c.execute("ALTER TABLE composizione ADD COLUMN unita TEXT")
    except sqlite3.OperationalError:
        pass  # Colonna esiste già
    
    conn.commit()
    conn.close()

def load_materie_prime():
    """Carica tutte le materie prime dal database"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT id, nome, costo, quantita, unita FROM materie_prime")
    data = c.fetchall()
    conn.close()
    return data

def get_materia_prima_by_id(materia_id):
    """Carica una materia prima specifica"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT id, nome, costo, quantita, unita FROM materie_prime WHERE id = ?", (materia_id,))
    data = c.fetchone()
    conn.close()
    return data

def save_materie_prime(table):
    """Salva tutte le materie prime nel database"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("DELETE FROM materie_prime")
    
    for row in range(table.rowCount()):
        nome = table.item(row, 0).text() if table.item(row, 0) else ""
        costo = parse_float(table.item(row, 1).text() if table.item(row, 1) else "0")
        quantita = parse_float(table.item(row, 2).text() if table.item(row, 2) else "0")
        combo = table.cellWidget(row, 3)
        unita = combo.currentText() if combo else ""
        
        if nome:
            c.execute("INSERT INTO materie_prime (nome, costo, quantita, unita) VALUES (?, ?, ?, ?)",
                     (nome, costo, quantita, unita))
    
    conn.commit()
    conn.close()

def load_prodotti():
    """Carica tutti i prodotti"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT id, nome, tipo FROM prodotti ORDER BY nome")
    data = c.fetchall()
    conn.close()
    return data

def get_prodotto_by_id(prodotto_id):
    """Carica un prodotto specifico"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT id, nome, tipo FROM prodotti WHERE id = ?", (prodotto_id,))
    data = c.fetchone()
    conn.close()
    return data

def save_prodotto(prodotto_id, nome, tipo):
    """Salva o aggiorna un prodotto"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    
    if prodotto_id is None:
        c.execute("INSERT INTO prodotti (nome, tipo) VALUES (?, ?)", (nome, tipo))
        conn.commit()
        new_id = c.lastrowid
        conn.close()
        return new_id
    else:
        c.execute("UPDATE prodotti SET nome = ?, tipo = ? WHERE id = ?", (nome, tipo, prodotto_id))
        conn.commit()
        conn.close()
        return prodotto_id

def delete_prodotto(prodotto_id):
    """Elimina un prodotto e la sua composizione"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("DELETE FROM composizione WHERE prodotto_id = ?", (prodotto_id,))
    c.execute("DELETE FROM prodotti WHERE id = ?", (prodotto_id,))
    conn.commit()
    conn.close()

def load_composizione(prodotto_id):
    """Carica la composizione di un prodotto"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT id, elemento_id, elemento_tipo, quantita, unita FROM composizione WHERE prodotto_id = ? ORDER BY id",
             (prodotto_id,))
    data = c.fetchall()
    conn.close()
    return data

def save_composizione(prodotto_id, composizione_data):
    """Salva la composizione di un prodotto"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("DELETE FROM composizione WHERE prodotto_id = ?", (prodotto_id,))
    
    for elemento_id, quantita, unita in composizione_data:
        if elemento_id:
            c.execute("INSERT INTO composizione (prodotto_id, elemento_id, elemento_tipo, quantita, unita) VALUES (?, ?, ?, ?, ?)",
                     (prodotto_id, elemento_id, "materia_prima", quantita, unita))
    
    conn.commit()
    conn.close()

def elimina_composizione_riga(composizione_id):
    """Elimina una riga di composizione"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("DELETE FROM composizione WHERE id = ?", (composizione_id,))
    conn.commit()
    conn.close()

def calcola_costo_prodotto(prodotto_id, visited=None, profondita=0):
    """Calcola il costo totale di un prodotto ricorsivamente"""
    if visited is None:
        visited = set()
    
    if profondita > 5:
        return 0.0
    
    if prodotto_id in visited:
        return 0.0
    
    visited.add(prodotto_id)
    
    costo_totale = 0.0
    composizione = load_composizione(prodotto_id)
    
    for comp_id, elemento_id, elemento_tipo, quantita, unita in composizione:
        quantita = parse_float(quantita)
        
        if elemento_tipo == "materia_prima":
            materia = get_materia_prima_by_id(elemento_id)
            if materia:
                id_m, nome, costo, qta, mat_unita = materia
                
                # Converti la quantità all'unità base della materia prima
                fattore_da = FATTORI_CONVERSIONE.get(unita, 1.0) if unita else FATTORI_CONVERSIONE.get(mat_unita, 1.0)
                fattore_a = FATTORI_CONVERSIONE.get(mat_unita, 1.0)
                quantita_base = quantita * fattore_a / fattore_da
                
                costo_unitario = costo / qta if qta > 0 else 0
                costo_totale += costo_unitario * quantita_base
        
        elif elemento_tipo == "prodotto":
            costo_sub = calcola_costo_prodotto(elemento_id, visited.copy(), profondita + 1)
            costo_totale += costo_sub * quantita
    
    return costo_totale

def esporta_excel(table):
    """Esporta la tabella in un file Excel"""
    file_path, _ = QFileDialog.getSaveFileName(None, "Salva come Excel", "", "Excel Files (*.xlsx)")
    if not file_path:
        return
    
    if not file_path.endswith('.xlsx'):
        file_path += '.xlsx'
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Materie Prime"
    
    ws.append([COLONNA_NOME, COLONNA_COSTO, COLONNA_QUANTITA, COLONNA_UNITA])
    
    for row in range(table.rowCount()):
        nome = table.item(row, 0).text() if table.item(row, 0) else ""
        costo = table.item(row, 1).text() if table.item(row, 1) else ""
        quantita = table.item(row, 2).text() if table.item(row, 2) else ""
        combo = table.cellWidget(row, 3)
        unita = combo.currentText() if combo else ""
        
        if nome:
            ws.append([nome, parse_float(costo), parse_float(quantita), unita])
    
    wb.save(file_path)
    QMessageBox.information(None, "Successo", f"Dati esportati in: {file_path}")

def importa_excel(table):
    """Importa dati da un file Excel"""
    file_path, _ = QFileDialog.getOpenFileName(None, "Apri file Excel", "", "Excel Files (*.xlsx)")
    if not file_path:
        return
    
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        
        table.setRowCount(0)
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 0):
            if row[0] is None:
                break
            
            table.insertRow(table.rowCount())
            current_row = table.rowCount() - 1
            
            table.setItem(current_row, 0, QTableWidgetItem(str(row[0]) if row[0] else ""))
            table.setItem(current_row, 1, QTableWidgetItem(str(parse_float(row[1])) if row[1] else ""))
            table.setItem(current_row, 2, QTableWidgetItem(str(parse_float(row[2])) if row[2] else ""))
            
            combo = QComboBox()
            combo.addItems(UNITA_MISURA)
            combo.blockSignals(True)
            if row[3] and str(row[3]) in UNITA_MISURA:
                combo.setCurrentText(str(row[3]))
            combo.blockSignals(False)
            combo.currentIndexChanged.connect(lambda: save_materie_prime(table))
            table.setCellWidget(current_row, 3, combo)
        
        save_materie_prime(table)
        QMessageBox.information(None, "Successo", "Dati importati correttamente")
    except Exception as e:
        QMessageBox.critical(None, "Errore", f"Errore nell'importazione: {str(e)}")

class NumericDelegate(QStyledItemDelegate):
    """Delegate che converte automaticamente virgole in punti durante la digitazione"""
    
    def createEditor(self, parent, option, index):
        editor = QLineEdit(parent)
        editor.textChanged.connect(lambda text: self._converti_virgola(editor))
        return editor
    
    def setEditorData(self, editor, index):
        value = index.model().data(index, Qt.EditRole)
        editor.setText(str(value) if value else "")
        editor.textChanged.connect(lambda text: self._converti_virgola(editor))
    
    def setModelData(self, editor, model, index):
        text = editor.text().replace(",", ".")
        model.setData(index, text, Qt.EditRole)
    
    def _converti_virgola(self, editor):
        cursor_pos = editor.cursorPosition()
        text = editor.text().replace(",", ".")
        editor.blockSignals(True)
        editor.setText(text)
        editor.blockSignals(False)
        editor.setCursorPosition(cursor_pos)

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle(WINDOW_TITLE)
        self.setGeometry(100, 100, WINDOW_WIDTH, WINDOW_HEIGHT)
        
        init_db()
        
        self.current_prodotto_id = None
        self.composizione_ids = {}
        
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        main_layout = QVBoxLayout()
        central_widget.setLayout(main_layout)
        
        tabs = QTabWidget()
        main_layout.addWidget(tabs)
        
        # Tab 1: Materie Prime
        tab1 = QWidget()
        tab1_layout = QVBoxLayout()
        tab1.setLayout(tab1_layout)
        
        self.table_materie = QTableWidget()
        self.setup_tabella_materie()
        tab1_layout.addWidget(self.table_materie)
        
        button_layout = QHBoxLayout()
        btn_add_row = QPushButton("Aggiungi Riga")
        btn_add_row.clicked.connect(self.aggiungi_riga_materie)
        btn_dup_mat = QPushButton("Duplica Materia Prima")
        btn_dup_mat.clicked.connect(self.duplica_materia_prima)
        button_export = QPushButton("Esporta in Excel")
        button_export.clicked.connect(lambda: esporta_excel(self.table_materie))
        button_import = QPushButton("Importa da Excel")
        button_import.clicked.connect(lambda: importa_excel(self.table_materie))
        button_layout.addWidget(btn_add_row)
        button_layout.addWidget(btn_dup_mat)
        button_layout.addWidget(button_export)
        button_layout.addWidget(button_import)
        tab1_layout.addLayout(button_layout)
        
        # Tab 2: Prodotti
        tab2 = QWidget()
        tab2_layout = QHBoxLayout()
        tab2.setLayout(tab2_layout)
        
        # Panel sinistro: tabella prodotti
        left_layout = QVBoxLayout()
        
        self.table_prodotti = QTableWidget()
        self.setup_tabella_prodotti()
        left_layout.addWidget(self.table_prodotti)
        
        button_prodotti_layout = QVBoxLayout()
        btn_add = QPushButton("Aggiungi Prodotto")
        btn_add.clicked.connect(self.aggiungi_prodotto)
        btn_del = QPushButton("Elimina Prodotto")
        btn_del.clicked.connect(self.elimina_prodotto)
        btn_dup = QPushButton("Duplica Prodotto")
        btn_dup.clicked.connect(self.duplica_prodotto)
        
        button_prodotti_layout.addWidget(btn_add)
        button_prodotti_layout.addWidget(btn_del)
        button_prodotti_layout.addWidget(btn_dup)
        left_layout.addLayout(button_prodotti_layout)
        
        # Panel destro: composizione
        right_layout = QVBoxLayout()
        right_layout.addWidget(QLabel("Composizione:"))
        
        self.table_composizione = QTableWidget()
        self.setup_tabella_composizione()
        right_layout.addWidget(self.table_composizione)
        
        btn_add_elem = QPushButton("Aggiungi Ingrediente")
        btn_add_elem.clicked.connect(self.aggiungi_ingrediente)
        btn_del_elem = QPushButton("Elimina Ingrediente")
        btn_del_elem.clicked.connect(self.elimina_ingrediente)
        
        right_layout.addWidget(btn_add_elem)
        right_layout.addWidget(btn_del_elem)
        
        left_widget = QWidget()
        left_widget.setLayout(left_layout)
        left_widget.setMinimumWidth(300)
        
        right_widget = QWidget()
        right_widget.setLayout(right_layout)
        
        tab2_layout.addWidget(left_widget, 1)
        tab2_layout.addWidget(right_widget, 1)
        
        # Tab 3: vuoto
        tab3 = QWidget()
        
        tabs.addTab(tab1, NOME_TAB_1)
        tabs.addTab(tab2, NOME_TAB_2)
        tabs.addTab(tab3, NOME_TAB_3)
        
        # Connetti selezione prodotto
        self.table_prodotti.itemSelectionChanged.connect(self.on_prodotto_selezionato)
    
    def setup_tabella_materie(self):
        """Configura la tabella delle materie prime"""
        self.table_materie.setColumnCount(4)
        self.table_materie.setHorizontalHeaderLabels([
            COLONNA_NOME,
            COLONNA_COSTO,
            COLONNA_QUANTITA,
            COLONNA_UNITA
        ])
        
        data = load_materie_prime()
        self.table_materie.setRowCount(len(data) if data else 5)
        
        numeric_delegate = NumericDelegate()
        self.table_materie.setItemDelegateForColumn(1, numeric_delegate)
        
        for row, record in enumerate(data) if data else enumerate([]):
            id_db, nome, costo, quantita, unita = record
            self.table_materie.setItem(row, 0, QTableWidgetItem(nome))
            self.table_materie.setItem(row, 1, QTableWidgetItem(str(costo)))
            self.table_materie.setItem(row, 2, QTableWidgetItem(str(quantita)))
            
            combo = QComboBox()
            combo.addItems(UNITA_MISURA)
            combo.blockSignals(True)
            combo.setCurrentText(unita)
            combo.blockSignals(False)
            combo.currentIndexChanged.connect(lambda: save_materie_prime(self.table_materie))
            self.table_materie.setCellWidget(row, 3, combo)
        
        if not data:
            for row in range(5):
                self.table_materie.setItem(row, 0, QTableWidgetItem(""))
                self.table_materie.setItem(row, 1, QTableWidgetItem(""))
                self.table_materie.setItem(row, 2, QTableWidgetItem(""))
                
                combo = QComboBox()
                combo.addItems(UNITA_MISURA)
                combo.currentIndexChanged.connect(lambda: save_materie_prime(self.table_materie))
                self.table_materie.setCellWidget(row, 3, combo)
        
        self.table_materie.setEditTriggers(QTableWidget.DoubleClicked | QTableWidget.SelectedClicked)
        self.table_materie.cellChanged.connect(lambda: save_materie_prime(self.table_materie))
        self.table_materie.resizeColumnsToContents()
    
    def aggiungi_riga_materie(self):
        """Aggiunge una riga vuota alla tabella materie"""
        row = self.table_materie.rowCount()
        self.table_materie.insertRow(row)
        
        self.table_materie.setItem(row, 0, QTableWidgetItem(""))
        self.table_materie.setItem(row, 1, QTableWidgetItem(""))
        self.table_materie.setItem(row, 2, QTableWidgetItem(""))
        
        combo = QComboBox()
        combo.addItems(UNITA_MISURA)
        combo.currentIndexChanged.connect(lambda: save_materie_prime(self.table_materie))
        self.table_materie.setCellWidget(row, 3, combo)
    
    def duplica_materia_prima(self):
        """Duplica la materia prima selezionata"""
        row = self.table_materie.currentRow()
        if row < 0:
            QMessageBox.warning(None, "Errore", "Seleziona una materia prima")
            return
        
        nome = self.table_materie.item(row, 0).text()
        costo = self.table_materie.item(row, 1).text()
        quantita = self.table_materie.item(row, 2).text()
        combo = self.table_materie.cellWidget(row, 3)
        unita = combo.currentText()
        
        if not nome:
            QMessageBox.warning(None, "Errore", "La riga non ha un nome")
            return
        
        self.aggiungi_riga_materie()
        new_row = self.table_materie.rowCount() - 1
        
        self.table_materie.setItem(new_row, 0, QTableWidgetItem(f"{nome} (copia)"))
        self.table_materie.setItem(new_row, 1, QTableWidgetItem(costo))
        self.table_materie.setItem(new_row, 2, QTableWidgetItem(quantita))
        combo_new = self.table_materie.cellWidget(new_row, 3)
        combo_new.setCurrentText(unita)
        
        save_materie_prime(self.table_materie)
    
    def setup_tabella_prodotti(self):
        """Configura la tabella dei prodotti"""
        self.table_prodotti.setColumnCount(2)
        self.table_prodotti.setHorizontalHeaderLabels(["Prodotto", "Costo Unitario (€)"])
        self.table_prodotti.setEditTriggers(QTableWidget.NoEditTriggers)
        self.carica_prodotti()
    
    def setup_tabella_composizione(self):
        """Configura la tabella di composizione"""
        self.table_composizione.setColumnCount(4)
        self.table_composizione.setHorizontalHeaderLabels(["Materia Prima", "Unità", "Quantità", "Costo Parziale"])
    
    def carica_prodotti(self):
        """Carica i prodotti nella tabella"""
        prodotti = load_prodotti()
        self.table_prodotti.setRowCount(len(prodotti))
        
        for row, (prod_id, nome, tipo) in enumerate(prodotti):
            costo = calcola_costo_prodotto(prod_id)
            
            item_nome = QTableWidgetItem(f"{nome} ({tipo})")
            item_nome.setData(Qt.UserRole, prod_id)
            item_nome.setFlags(item_nome.flags() & ~Qt.ItemIsEditable)
            self.table_prodotti.setItem(row, 0, item_nome)
            
            item_costo = QTableWidgetItem(f"{costo:.2f}")
            item_costo.setFlags(item_costo.flags() & ~Qt.ItemIsEditable)
            self.table_prodotti.setItem(row, 1, item_costo)
        
        self.table_prodotti.resizeColumnsToContents()
    
    def on_prodotto_selezionato(self):
        """Quando si seleziona un prodotto, carica la composizione"""
        row = self.table_prodotti.currentRow()
        if row < 0:
            self.current_prodotto_id = None
            self.table_composizione.setRowCount(0)
            return
        
        item = self.table_prodotti.item(row, 0)
        self.current_prodotto_id = item.data(Qt.UserRole)
        self.carica_composizione()
    
    def carica_composizione(self):
        """Carica la composizione del prodotto selezionato"""
        if not self.current_prodotto_id:
            return
        
        self.table_composizione.blockSignals(True)
        self.table_composizione.setRowCount(0)
        self.composizione_ids = {}
        
        composizione = load_composizione(self.current_prodotto_id)
        materie = load_materie_prime()
        
        for comp_id, elemento_id, elemento_tipo, quantita, unita in composizione:
            row = self.table_composizione.rowCount()
            self.table_composizione.insertRow(row)
            self.composizione_ids[row] = comp_id
            
            # Colonna 0: Materia Prima (combobox)
            combo_materia = QComboBox()
            combo_materia.addItem("-- Seleziona --", None)
            for mat_id, nome, costo, qta, mat_unita in materie:
                combo_materia.addItem(nome, mat_id)
            
            if elemento_id:
                for i in range(combo_materia.count()):
                    if combo_materia.itemData(i) == elemento_id:
                        combo_materia.setCurrentIndex(i)
                        break
            
            combo_materia.currentIndexChanged.connect(self.on_cambio_materia_prima)
            self.table_composizione.setCellWidget(row, 0, combo_materia)
            
            # Colonna 1: Unità di Misura (combobox con unità compatibili)
            combo_unita = QComboBox()
            self.popola_unita_compatibili(combo_unita, elemento_id)
            if unita:
                combo_unita.blockSignals(True)
                combo_unita.setCurrentText(unita)
                combo_unita.blockSignals(False)
            combo_unita.currentIndexChanged.connect(self.on_cambio_unita)
            self.table_composizione.setCellWidget(row, 1, combo_unita)
            
            # Colonna 2: Quantità
            spin_quantita = QDoubleSpinBox()
            spin_quantita.setValue(parse_float(quantita))
            spin_quantita.setMinimum(0)
            spin_quantita.setMaximum(10000)
            spin_quantita.setSingleStep(0.1)
            spin_quantita.setObjectName(f"spin_{row}")
            spin_quantita.valueChanged.connect(self.on_cambio_quantita)
            self.table_composizione.setCellWidget(row, 2, spin_quantita)
            
            # Colonna 3: Costo Parziale (read-only)
            self.aggiorna_costo_parziale(row, elemento_id, quantita, unita)
        
        self.table_composizione.blockSignals(False)
    
    def popola_unita_compatibili(self, combo_unita, materia_id):
        """Popola il combobox unità con solo le unità compatibili della materia prima"""
        combo_unita.clear()
        combo_unita.addItem("-- Seleziona --", None)
        
        if materia_id:
            materia = get_materia_prima_by_id(materia_id)
            if materia:
                id_m, nome, costo, qta, mat_unita = materia
                
                # Unità compatibili per tipo
                if mat_unita in ["kg", "grammi"]:
                    for u in ["kg", "grammi"]:
                        combo_unita.addItem(u, u)
                elif mat_unita in ["litri", "ml"]:
                    for u in ["litri", "ml"]:
                        combo_unita.addItem(u, u)
                elif mat_unita in ["m", "cm"]:
                    for u in ["m", "cm"]:
                        combo_unita.addItem(u, u)
                else:
                    combo_unita.addItem(mat_unita, mat_unita)
    
    def on_cambio_materia_prima(self):
        """Quando cambia la materia prima, aggiorna le unità disponibili"""
        row = self.table_composizione.currentRow()
        if row < 0:
            return
        
        combo_materia = self.table_composizione.cellWidget(row, 0)
        combo_unita = self.table_composizione.cellWidget(row, 1)
        
        if combo_materia and combo_unita:
            materia_id = combo_materia.currentData()
            self.popola_unita_compatibili(combo_unita, materia_id)
            self.on_cambio_quantita()
    
    def on_cambio_unita(self):
        """Quando cambia l'unità"""
        self.on_cambio_quantita()
    
    def on_cambio_quantita(self):
        """Quando cambia la quantità, aggiorna il costo"""
        # Trova la riga che ha triggerato il segnale
        for row in range(self.table_composizione.rowCount()):
            combo_materia = self.table_composizione.cellWidget(row, 0)
            combo_unita = self.table_composizione.cellWidget(row, 1)
            spin_quantita = self.table_composizione.cellWidget(row, 2)
            
            if spin_quantita:
                materia_id = combo_materia.currentData() if combo_materia else None
                unita = combo_unita.currentText() if combo_unita else ""
                quantita = spin_quantita.value()
                
                self.aggiorna_costo_parziale(row, materia_id, quantita, unita)
        
        self.salva_composizione()
        self.aggiorna_costo_prodotto()
    
    def aggiorna_costo_parziale(self, row, materia_id, quantita, unita):
        """Aggiorna il costo parziale di una riga"""
        costo_parziale = 0.0
        
        if materia_id and unita:
            materia = get_materia_prima_by_id(materia_id)
            if materia:
                id_m, nome, costo, qta, mat_unita = materia
                
                # Converti la quantità all'unità base della materia prima
                fattore_da = FATTORI_CONVERSIONE.get(unita, 1.0)
                fattore_a = FATTORI_CONVERSIONE.get(mat_unita, 1.0)
                quantita_base = quantita * fattore_a / fattore_da
                
                costo_unitario = costo / qta if qta > 0 else 0
                costo_parziale = costo_unitario * quantita_base
        
        item = QTableWidgetItem(f"{costo_parziale:.2f}")
        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
        item.setBackground(QBrush(QColor(200, 200, 200)))
        self.table_composizione.setItem(row, 3, item)
    
    def salva_composizione(self):
        """Salva la composizione nel database"""
        if not self.current_prodotto_id:
            return
        
        composizione_data = []
        for row in range(self.table_composizione.rowCount()):
            combo_materia = self.table_composizione.cellWidget(row, 0)
            combo_unita = self.table_composizione.cellWidget(row, 1)
            spin_quantita = self.table_composizione.cellWidget(row, 2)
            
            if combo_materia and spin_quantita:
                materia_id = combo_materia.currentData()
                unita = combo_unita.currentText() if combo_unita else ""
                quantita = spin_quantita.value()
                
                if materia_id and unita != "-- Seleziona --":
                    composizione_data.append((materia_id, quantita, unita))
        
        save_composizione(self.current_prodotto_id, composizione_data)
    
    def aggiorna_costo_prodotto(self):
        """Aggiorna il costo del prodotto nella tabella sinistra"""
        if not self.current_prodotto_id:
            return
        
        costo = calcola_costo_prodotto(self.current_prodotto_id)
        
        for row in range(self.table_prodotti.rowCount()):
            item = self.table_prodotti.item(row, 0)
            if item.data(Qt.UserRole) == self.current_prodotto_id:
                item_costo = self.table_prodotti.item(row, 1)
                item_costo.setText(f"{costo:.2f}")
                break
    
    def aggiungi_ingrediente(self):
        """Aggiunge una riga alla composizione"""
        if not self.current_prodotto_id:
            QMessageBox.warning(None, "Errore", "Seleziona un prodotto")
            return
        
        row = self.table_composizione.rowCount()
        self.table_composizione.insertRow(row)
        self.composizione_ids[row] = None
        
        # Colonna 0: Materia Prima
        combo_materia = QComboBox()
        combo_materia.addItem("-- Seleziona --", None)
        for mat_id, nome, costo, qta, unita in load_materie_prime():
            combo_materia.addItem(nome, mat_id)
        
        combo_materia.currentIndexChanged.connect(self.on_cambio_materia_prima)
        self.table_composizione.setCellWidget(row, 0, combo_materia)
        
        # Colonna 1: Unità
        combo_unita = QComboBox()
        combo_unita.addItem("-- Seleziona --", None)
        combo_unita.currentIndexChanged.connect(self.on_cambio_unita)
        self.table_composizione.setCellWidget(row, 1, combo_unita)
        
        # Colonna 2: Quantità
        spin_quantita = QDoubleSpinBox()
        spin_quantita.setMinimum(0)
        spin_quantita.setMaximum(10000)
        spin_quantita.setSingleStep(0.1)
        spin_quantita.valueChanged.connect(self.on_cambio_quantita)
        self.table_composizione.setCellWidget(row, 2, spin_quantita)
        
        # Colonna 3: Costo Parziale
        item = QTableWidgetItem("0.00")
        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
        item.setBackground(QBrush(QColor(200, 200, 200)))
        self.table_composizione.setItem(row, 3, item)
    
    def elimina_ingrediente(self):
        """Elimina un ingrediente dalla composizione"""
        row = self.table_composizione.currentRow()
        if row < 0:
            QMessageBox.warning(None, "Errore", "Seleziona un ingrediente")
            return
        
        comp_id = self.composizione_ids.get(row)
        if comp_id:
            elimina_composizione_riga(comp_id)
        
        self.table_composizione.removeRow(row)
        self.salva_composizione()
        self.aggiorna_costo_prodotto()
    
    def aggiungi_prodotto(self):
        """Aggiunge un nuovo prodotto"""
        save_prodotto(None, "Nuovo Prodotto", TIPI_PRODOTTO[0])
        self.carica_prodotti()
    
    def elimina_prodotto(self):
        """Elimina il prodotto selezionato"""
        row = self.table_prodotti.currentRow()
        if row < 0:
            QMessageBox.warning(None, "Errore", "Seleziona un prodotto")
            return
        
        item = self.table_prodotti.item(row, 0)
        prod_id = item.data(Qt.UserRole)
        
        if QMessageBox.question(None, "Conferma", "Eliminare questo prodotto?") == QMessageBox.Yes:
            delete_prodotto(prod_id)
            self.carica_prodotti()
            self.current_prodotto_id = None
            self.table_composizione.setRowCount(0)
    
    def duplica_prodotto(self):
        """Duplica il prodotto selezionato"""
        row = self.table_prodotti.currentRow()
        if row < 0:
            QMessageBox.warning(None, "Errore", "Seleziona un prodotto")
            return
        
        item = self.table_prodotti.item(row, 0)
        prod_id = item.data(Qt.UserRole)
        
        prodotto = get_prodotto_by_id(prod_id)
        if not prodotto:
            return
        
        id_p, nome, tipo = prodotto
        new_nome = f"{nome} (copia)"
        new_id = save_prodotto(None, new_nome, tipo)
        
        composizione = load_composizione(prod_id)
        comp_data = [(el_id, qta, unita) for _, el_id, el_tipo, qta, unita in composizione]
        save_composizione(new_id, comp_data)
        
        self.carica_prodotti()

# ============================================================================
# MAIN
# ============================================================================

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())